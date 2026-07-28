#!/usr/bin/env python3
"""Cổng QA cho file .xlsx — chặn publish khi công thức sai.

    .venv/bin/python scripts/qa_check.py

Spec do AI soạn nên công thức có thể sai một cách âm thầm: file vẫn mở được,
vẫn ra số, nhưng số đó sai. Script này tính lại toàn bộ workbook bằng thư viện
`formulas` (không cần cài Excel) và kiểm tra ba nhóm rủi ro:

  1. Ô báo lỗi Excel (#DIV/0!, #REF!, #N/A, #VALUE!, #NAME?, #NUM!, #NULL!)
  2. Công thức chết khi người dùng nhập thêm dữ liệu vào dòng trống
  3. Tham chiếu tuyệt đối ($) đặt nhầm, làm công thức sai khi kéo xuống

Vẫn phải mở bằng Excel thật một lần trước khi publish — script này bắt lỗi máy
kiểm được, không thay thế được mắt người soát ý nghĩa nghiệp vụ.
"""

from __future__ import annotations

import json
import re
import shutil
import sys
import tempfile
import warnings
from datetime import date
from pathlib import Path

import formulas
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import build_bundle

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data" / "templates"
SYSTEMS_DIR = ROOT / "data" / "systems"
OUT_DIR = ROOT / "public" / "downloads"
COMPUTED_DIR = ROOT / "data" / "computed"

ERROR_VALUES = {
    "#DIV/0!",
    "#REF!",
    "#N/A",
    "#VALUE!",
    "#NAME?",
    "#NUM!",
    "#NULL!",
}

# Khóa của thư viện `formulas` có dạng "'[FILE.XLSX]TÊN SHEET'!A5", tên sheet
# đã viết hoa. Phải bắt cả tên sheet: bỏ nó đi thì ô A2 của sheet này ghi đè ô
# A2 của sheet kia, và file nhiều sheet được kiểm sót một cách âm thầm.
CELL_RE = re.compile(r"^'\[[^\]]+\](.+)'!([A-Z]+)(\d+)$")


def cell_value(raw):
    """Ranges object của `formulas` → giá trị Python thuần."""
    try:
        value = raw.value[0, 0]
    except (AttributeError, IndexError, TypeError):
        return raw
    return value.item() if hasattr(value, "item") else value


def evaluate(path: Path) -> dict[tuple[str, str, int], object]:
    """Tính lại workbook, trả về {(TÊN SHEET, cột, dòng): giá trị}."""
    model = formulas.ExcelModel().loads(str(path)).finish()
    solution = model.calculate()

    values: dict[tuple[str, str, int], object] = {}
    for key, raw in solution.items():
        match = CELL_RE.match(key)
        if not match:
            continue
        sheet, col, row = match.group(1), match.group(2), int(match.group(3))
        values[(sheet.upper(), col, row)] = cell_value(raw)
    return values


def find_errors(values: dict[tuple[str, str, int], object]) -> list[str]:
    problems = []
    for (sheet, col, row), value in sorted(values.items()):
        text = str(value).strip()
        if text in ERROR_VALUES:
            problems.append(f"ô {col}{row} sheet \"{sheet}\" báo lỗi {text}")
    return problems


def check_absolute_refs(spec: dict) -> list[str]:
    """Tham chiếu $ trong công thức theo dòng sẽ trỏ sai khi kéo xuống."""
    problems = []
    for sheet in spec["sheets"]:
        for col in sheet["columns"]:
            formula = col.get("formula")
            if formula and "$" in formula:
                problems.append(
                    f"cột \"{col['header']}\" dùng tham chiếu tuyệt đối ($) — "
                    f"công thức sẽ trỏ sai dòng khi người dùng kéo xuống: {formula}"
                )
    return problems


def fill_blank_rows(src: Path, spec: dict, count: int = 5) -> Path:
    """Nhập dữ liệu giả vào các dòng trống, mô phỏng người dùng dùng thật."""
    tmp = Path(tempfile.mkdtemp()) / src.name
    shutil.copy(src, tmp)

    wb = load_workbook(tmp)
    for sheet in spec["sheets"]:
        ws = wb[sheet["name"]]
        first_blank = 2 + len(sheet["sampleRows"])

        for offset in range(count):
            row = first_blank + offset
            for i, col in enumerate(sheet["columns"], start=1):
                if col["type"] == "formula":
                    continue
                ws.cell(row=row, column=i, value=fake_value(col, offset))

    wb.save(tmp)
    return tmp


def fake_value(col: dict, offset: int):
    """Giá trị hợp lệ theo kiểu cột, tôn trọng ràng buộc validation."""
    rule = col.get("validation") or {}
    if col["type"] == "text":
        if rule.get("type") == "list" and rule.get("options"):
            return rule["options"][offset % len(rule["options"])]
        return f"QA {offset + 1}"
    if col["type"] == "date":
        return date(2026, 7, 1)
    if col["type"] in {"number", "currency", "percent"}:
        low = rule.get("min", 1)
        high = rule.get("max", 100)
        # Chọn giá trị nằm giữa khoảng cho phép để không đụng trần validation.
        return round(low + (high - low) * 0.5)
    return None


def export_computed(spec: dict, values: dict[tuple[str, int], object]) -> None:
    """Ghi giá trị các ô công thức của dòng mẫu ra cho website dùng.

    Trang preview hiển thị số thật thay vì chữ "tự tính", và vì file này chỉ
    được ghi sau khi QA đạt nên website không thể hiển thị một con số chưa qua
    kiểm tra.
    """
    sheets = []
    for sheet in spec["sheets"]:
        rows = []
        for row_index in range(len(sheet["sampleRows"])):
            row_num = 2 + row_index
            row: dict[str, object] = {}
            for i, col in enumerate(sheet["columns"], start=1):
                if col["type"] != "formula":
                    continue
                value = values.get((sheet["name"].upper(), get_column_letter(i), row_num))
                if value is None or str(value).strip() in {"", "empty"}:
                    continue
                row[col["key"]] = value
            rows.append(row)
        sheets.append(rows)

    out = COMPUTED_DIR / spec["category"] / f"{spec['slug']}.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        json.dumps(sheets, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def check_template(spec_path: Path) -> list[str]:
    spec = json.loads(spec_path.read_text(encoding="utf-8"))
    xlsx = OUT_DIR / spec["category"] / f"{spec['slug']}.xlsx"

    if not xlsx.exists():
        return [f"thiếu file {xlsx.relative_to(ROOT)} — chạy build_xlsx.py trước"]

    problems = check_absolute_refs(spec)
    original_values = evaluate(xlsx)
    problems += [f"[nguyên bản] {p}" for p in find_errors(original_values)]

    # Kiểm tra kịch bản thật: người dùng nhập thêm 5 dòng vào vùng trống.
    filled = fill_blank_rows(xlsx, spec)
    filled_values = evaluate(filled)
    problems += [f"[sau khi thêm 5 dòng] {p}" for p in find_errors(filled_values)]

    # Công thức phải thực sự chạy ra kết quả, không được im lặng trả về rỗng.
    for sheet in spec["sheets"]:
        first_blank = 2 + len(sheet["sampleRows"])
        for i, col in enumerate(sheet["columns"], start=1):
            if col["type"] != "formula":
                continue
            letter = get_column_letter(i)
            for offset in range(5):
                row = first_blank + offset
                value = filled_values.get((sheet["name"].upper(), letter, row))
                if value is None or str(value).strip() in {"", "empty"}:
                    problems.append(
                        f"[sau khi thêm 5 dòng] sheet \"{sheet['name']}\" cột "
                        f"\"{col['header']}\" tại {letter}{row} không ra kết quả "
                        "dù dòng đã có dữ liệu"
                    )
                    break

    # Chỉ xuất giá trị cho website khi mọi kiểm tra đều sạch.
    if not problems:
        export_computed(spec, original_values)

    return problems


def column_letter_of(columns: list[dict], key: str) -> str:
    return get_column_letter(next(i for i, c in enumerate(columns) if c["key"] == key) + 1)


def is_empty(value) -> bool:
    return value is None or str(value).strip() in {"", "empty"}


def check_links_join(bundle: build_bundle.Bundle, values: dict) -> list[str]:
    """Mỗi ô nối phải thực sự tra được dữ liệu trên các dòng mẫu.

    Đây là chỗ bắt lỗi nguy hiểm nhất của file gộp: công thức đúng cú pháp, file
    mở bình thường, không ô nào báo lỗi — nhưng khóa nối hai bên lệch nhau nên
    INDEX/MATCH và SUMIF không khớp được dòng nào và im lặng trả về rỗng hoặc 0.
    """
    problems = []

    for link in bundle.resolved:
        sheet = link["sheet"]
        columns = bundle.sheet_columns[sheet]
        letter = column_letter_of(columns, link["column"])
        sample_count = len(bundle.sources[sheet]["sampleRows"])

        found = False
        for row in range(2, 2 + sample_count):
            value = values.get((sheet.upper(), letter, row))
            if is_empty(value):
                problems.append(
                    f"ô nối \"{sheet}.{link['header']}\" tại {letter}{row} không ra kết quả"
                )
                break
            if isinstance(value, str) or value != 0:
                found = True
        else:
            if not found:
                problems.append(
                    f"ô nối \"{sheet}.{link['header']}\" trả về 0 ở mọi dòng mẫu — "
                    f"nhiều khả năng khóa {bundle.spec['keyName']} hai bên không khớp nhau"
                )

    return problems


def check_propagation(bundle: build_bundle.Bundle, xlsx: Path) -> list[str]:
    """Sửa một ô đầu vào thì sheet tổng phải đổi theo.

    Cả lớp bộ file dựa trên đúng lời hứa này. Ba kiểm tra ở trên chỉ nói công
    thức chạy được; chỉ có phép thử này nói dữ liệu thật sự chảy từ đầu này
    sang đầu kia của workbook.
    """
    masters = [n for n, role in bundle.role_by_sheet.items() if role == "master"]
    if not masters:
        return ["bộ có file gộp nhưng không sheet nào mang vai trò tổng hợp"]

    # Cột nhập liệu dạng số mà một công thức nối nào đó đọc tới — sửa vào đây là
    # tác động đúng vào đường dẫn dữ liệu mà bộ tuyên bố có.
    targets: list[tuple[str, str, dict]] = []
    for link in bundle.spec["links"]:
        for ref in re.findall(r"\[([^\]]+)\]", link["formula"]):
            if "!" not in ref:
                continue
            sheet, key = ref.split("!", 1)
            if bundle.role_by_sheet.get(sheet) == "master":
                continue
            column = next(c for c in bundle.sheet_columns[sheet] if c["key"] == key)
            if column["type"] not in {"number", "currency"}:
                continue
            entry = (sheet, column_letter_of(bundle.sheet_columns[sheet], key), column)
            if entry not in targets:
                targets.append(entry)

    if not targets:
        return ["không có cột số nào ở sheet đầu vào được công thức nối đọc tới"]

    before = evaluate(xlsx)

    tmp = Path(tempfile.mkdtemp()) / xlsx.name
    shutil.copy(xlsx, tmp)
    wb = load_workbook(tmp)
    for sheet, letter, column in targets:
        cell = wb[sheet][f"{letter}2"]
        delta = 1 if column["type"] == "number" else 100000
        cell.value = (cell.value or 0) + delta
    wb.save(tmp)

    after = evaluate(tmp)

    for master in masters:
        formula_keys = [
            c for c in bundle.sheet_columns[master] if c["type"] == "formula"
        ]
        for column in formula_keys:
            letter = column_letter_of(bundle.sheet_columns[master], column["key"])
            for row in range(2, 2 + len(bundle.sources[master]["sampleRows"])):
                key = (master.upper(), letter, row)
                if before.get(key) != after.get(key):
                    return []

    changed = ", ".join(f"{s}.{c['header']}" for s, _, c in targets)
    return [
        f"sửa {changed} ở dòng 2 nhưng không ô nào trên sheet tổng đổi theo — "
        "chuỗi liên kết đang đứt ở đâu đó"
    ]


def check_bundle(system_path: Path) -> list[str] | None:
    """None = bộ này chưa khai bundle, không có gì để kiểm."""
    bundle = build_bundle.load(system_path)
    if bundle is None:
        return None

    xlsx = bundle.out_path
    if not xlsx.exists():
        return [f"thiếu file {xlsx.relative_to(ROOT)} — chạy build_bundle.py trước"]

    values = evaluate(xlsx)
    problems = [f"[nguyên bản] {p}" for p in find_errors(values)]
    problems += check_links_join(bundle, values)
    problems += check_propagation(bundle, xlsx)
    return problems


def main() -> int:
    specs = sorted(DATA_DIR.glob("*/*.json"))
    if not specs:
        print("Không tìm thấy spec nào.", file=sys.stderr)
        return 1

    failed = 0
    for spec_path in specs:
        problems = check_template(spec_path)
        if problems:
            failed += 1
            print(f"\n✗ {spec_path.stem}")
            for problem in problems:
                print(f"    {problem}")
        else:
            print(f"✓ {spec_path.stem}")

    bundles = 0
    for system_path in sorted(SYSTEMS_DIR.glob("*.json")):
        problems = check_bundle(system_path)
        if problems is None:
            continue
        bundles += 1
        if problems:
            failed += 1
            print(f"\n✗ file gộp {system_path.stem}")
            for problem in problems:
                print(f"    {problem}")
        else:
            print(f"✓ file gộp {system_path.stem}")

    total = len(specs) + bundles
    if failed:
        print(f"\n{failed}/{total} file KHÔNG đạt QA — chưa được publish.")
        return 1

    print(f"\n{total}/{total} file đạt QA tự động.")
    print("Còn một bước bắt buộc: mở bằng Excel thật để soát ý nghĩa nghiệp vụ.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
