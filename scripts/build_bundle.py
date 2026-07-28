#!/usr/bin/env python3
"""Sinh file gộp của mỗi bộ — cả quy trình nằm trong một workbook.

    .venv/bin/python scripts/build_bundle.py

Mỗi file trong bộ góp một sheet, và các cột nối được đổi từ ô nhập liệu thành
công thức trỏ sang sheet khác. Nhờ nằm chung một workbook, liên kết không bao
giờ trả về #REF! vì đổi tên file hay chuyển thư mục — thứ luôn xảy ra khi nối
các file .xlsx rời với nhau.

resolve_bundle_formula dưới đây phải cho ra kết quả giống hệt
resolveBundleFormula trong lib/systems-schema.ts, nếu không công thức in trên
trang sẽ khác công thức nằm trong file người ta tải về.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from build_xlsx import SITE_URL, TITLE_FONT, build_sheet

ROOT = Path(__file__).resolve().parent.parent
SYSTEMS_DIR = ROOT / "data" / "systems"
TEMPLATES_DIR = ROOT / "data" / "templates"
OUT_DIR = ROOT / "public" / "downloads" / "bo-file"

# Phải khớp BUNDLE_ROW_LIMIT và BUNDLE_GUIDE_SHEET trong lib/systems-schema.ts.
ROW_LIMIT = 500
GUIDE_SHEET = "Hướng dẫn"

ROLE_LABEL = {"input": "Đầu vào", "process": "Xử lý", "master": "Tổng hợp"}


def resolve_bundle_formula(
    formula: str,
    host_columns: list[dict],
    sheet_columns: dict[str, list[dict]],
) -> str:
    """`[key]` → ô cùng sheet, `[Tên sheet!key]` → vùng cột tuyệt đối sheet khác.

    Giữ nguyên `{row}`; build_sheet mới là chỗ thay nó theo từng dòng.
    """
    host_letters = {c["key"]: get_column_letter(i + 1) for i, c in enumerate(host_columns)}

    def sub(match: re.Match[str]) -> str:
        ref = match.group(1)
        if "!" not in ref:
            if ref not in host_letters:
                raise KeyError(f"công thức nối trỏ tới cột không có: [{ref}]")
            return host_letters[ref]

        sheet_name, key = ref.split("!", 1)
        columns = sheet_columns.get(sheet_name)
        if columns is None:
            raise KeyError(f"công thức nối trỏ tới sheet không có trong bộ: [{ref}]")

        index = next((i for i, c in enumerate(columns) if c["key"] == key), None)
        if index is None:
            raise KeyError(f"sheet \"{sheet_name}\" không có cột nào dùng key \"{key}\"")

        letter = get_column_letter(index + 1)
        # Tuyệt đối cả cột lẫn dòng: vùng tra cứu phải đứng yên khi kéo xuống.
        return f"'{sheet_name}'!${letter}$2:${letter}${ROW_LIMIT}"

    return re.sub(r"\[([^\]]+)\]", sub, formula)


def build_guide_sheet(wb: Workbook, system: dict, links: list[dict]) -> None:
    """Sheet hướng dẫn của file gộp.

    Khác sheet hướng dẫn của file lẻ ở chỗ nó phải trả lời được hai câu mà chỉ
    file gộp mới đặt ra: nhập sheet nào trước, và ô nào không được gõ đè lên.
    """
    ws = wb.create_sheet(GUIDE_SHEET)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100

    bundle = system["bundle"]
    node_by_slug = {n["slug"]: n for n in system["nodes"]}
    heading = Font(bold=True, size=12)

    rows: list[tuple[str, Font | None]] = [
        (system["h1"], TITLE_FONT),
        (bundle["summary"], None),
        ("", None),
        ("THỨ TỰ NHẬP", heading),
        (
            "Nhập theo thứ tự tab từ trái sang phải. Cột nền xanh nhạt là ô có công thức — "
            "gõ đè lên sẽ xóa mất công thức của riêng ô đó.",
            None,
        ),
    ]

    for i, sheet in enumerate(bundle["sheets"], start=1):
        node = node_by_slug[sheet["node"]]
        rows.append((f"{i}. {sheet['name']}  —  {ROLE_LABEL[node['role']]}", Font(bold=True)))
        rows.append((f"   {node['owner']}.", None))
    rows.append(("", None))

    rows.append(("CÔNG THỨC NỐI GIỮA CÁC SHEET", heading))
    rows.append(
        (
            f"Đây là phần mà bản tải từng file riêng không có. Khóa nối giữa các sheet là "
            f"{bundle['keyName']}.",
            None,
        )
    )
    for link in links:
        rows.append(
            (f"{link['sheet']} · {link['header']}  —  {link['resolved']}", Font(bold=True))
        )
        rows.append((f"   {link['note']}", None))
    rows.append(("", None))

    rows.append(("GIỚI HẠN CẦN BIẾT", heading))
    rows.append(
        (
            f"Công thức nối quét tới dòng {ROW_LIMIT} của mỗi sheet. Dữ liệu vượt quá dòng đó "
            f"sẽ không được cộng vào các sheet sau mà không có cảnh báo nào. Nếu công ty bạn "
            f"cần nhiều hơn, sửa số {ROW_LIMIT} trong công thức lên mức cao hơn ở toàn bộ file.",
            None,
        )
    )
    rows.append(
        (
            "Cả bộ nằm trong một file nên chỉ nên có một người nhập tại một thời điểm. "
            "Muốn nhiều người nhập cùng lúc thì đưa file lên OneDrive hoặc SharePoint.",
            None,
        )
    )
    rows.append(("", None))

    url = f"{SITE_URL}/mau-excel/bo-file/{system['slug']}"
    rows.append(("Bản cập nhật mới nhất và sơ đồ liên kết:", Font(bold=True)))
    rows.append((url, Font(color="1D4ED8", underline="single")))

    for idx, (text, font) in enumerate(rows, start=1):
        cell = ws.cell(row=idx, column=2, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if font:
            cell.font = font

    ws.cell(row=len(rows), column=2).hyperlink = url


class Bundle:
    """Khối bundle đã ghép với template thật — dùng chung cho builder và cổng QA."""

    def __init__(self, system: dict) -> None:
        self.system = system
        self.spec = system["bundle"]

        specs: dict[str, dict] = {}
        for node in system["nodes"]:
            path = TEMPLATES_DIR / system["category"] / f"{node['slug']}.json"
            specs[node["slug"]] = json.loads(path.read_text(encoding="utf-8"))

        # Tên sheet trong file gộp → sheet spec gốc của template góp nó.
        self.sources = {
            s["name"]: specs[s["node"]]["sheets"][s.get("sheet", 0)]
            for s in self.spec["sheets"]
        }
        self.sheet_columns = {name: s["columns"] for name, s in self.sources.items()}
        self.role_by_sheet = {
            s["name"]: next(
                n["role"] for n in system["nodes"] if n["slug"] == s["node"]
            )
            for s in self.spec["sheets"]
        }

        self.resolved: list[dict] = []
        self.links_by_sheet: dict[str, dict[str, str]] = {}
        for link in self.spec["links"]:
            host_columns = self.sheet_columns[link["sheet"]]
            formula = resolve_bundle_formula(
                link["formula"], host_columns, self.sheet_columns
            )
            self.links_by_sheet.setdefault(link["sheet"], {})[link["column"]] = formula

            header = next(
                c["header"] for c in host_columns if c["key"] == link["column"]
            )
            self.resolved.append(
                {**link, "header": header, "resolved": formula.replace("{row}", "2")}
            )

    @property
    def out_path(self) -> Path:
        return OUT_DIR / f"{self.system['slug']}.xlsx"


def load(system_path: Path) -> Bundle | None:
    system = json.loads(system_path.read_text(encoding="utf-8"))
    return Bundle(system) if system.get("bundle") else None


def build(system_path: Path) -> Path | None:
    loaded = load(system_path)
    if loaded is None:
        return None

    system, bundle = loaded.system, loaded.spec
    sources, links_by_sheet, resolved = (
        loaded.sources,
        loaded.links_by_sheet,
        loaded.resolved,
    )

    wb = Workbook()
    for i, sheet in enumerate(bundle["sheets"]):
        build_sheet(
            wb,
            sources[sheet["name"]],
            first=(i == 0),
            title=sheet["name"],
            links=links_by_sheet.get(sheet["name"]),
        )
    build_guide_sheet(wb, system, resolved)

    out_path = loaded.out_path
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> int:
    systems = sorted(SYSTEMS_DIR.glob("*.json"))
    if not systems:
        print("Không tìm thấy bộ file nào trong data/systems/", file=sys.stderr)
        return 0

    built = 0
    for system_path in systems:
        try:
            out = build(system_path)
        except Exception as exc:  # noqa: BLE001
            print(f"✗ {system_path.name}: {exc}", file=sys.stderr)
            return 1
        if out is None:
            print(f"– {system_path.stem}: chưa khai bundle, bỏ qua")
            continue
        built += 1
        print(f"✓ {out.relative_to(ROOT)}  ({out.stat().st_size:,} bytes)")

    print(f"\nĐã sinh {built} file gộp.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
