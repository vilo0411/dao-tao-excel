#!/usr/bin/env python3
"""Sinh file .xlsx từ spec JSON — cùng nguồn dữ liệu mà trang web dùng.

    .venv/bin/python scripts/build_xlsx.py

Công thức trong spec viết theo dạng `[tenCot]{row}`; hàm resolve_formula dưới
đây phải cho ra kết quả giống hệt resolveFormula trong lib/templates.ts, nếu
không công thức hiển thị trên trang sẽ khác công thức trong file tải về.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data" / "templates"
OUT_DIR = ROOT / "public" / "downloads"

SITE_URL = "https://excel.nguyenvietloc.com"

HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
FORMULA_FILL = PatternFill("solid", fgColor="EFF6FF")
TITLE_FONT = Font(bold=True, size=14)

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUMBER_FORMATS = {
    "text": "@",
    "number": "#,##0.##",
    "date": "dd/mm/yyyy",
    "currency": '#,##0 "₫"',
    "percent": "0.0%",
}


def resolve_formula(formula: str, columns: list[dict], row: int) -> str:
    """`[key]{row}` → tham chiếu ô Excel thật. Song song với lib/templates.ts."""
    letters = {col["key"]: get_column_letter(i + 1) for i, col in enumerate(columns)}

    def sub(match: re.Match[str]) -> str:
        key = match.group(1)
        if key not in letters:
            raise KeyError(f"công thức trỏ tới cột không tồn tại: [{key}]")
        return letters[key]

    return re.sub(r"\[([^\]]+)\]", sub, formula).replace("{row}", str(row))


def apply_validation(ws, col: dict, letter: str, first_row: int, last_row: int) -> None:
    rule = col.get("validation")
    if not rule:
        return

    kind = rule["type"]
    if kind == "list":
        options = rule.get("options", [])
        # openpyxl nối options bằng dấu phẩy, nên option chứa dấu phẩy sẽ bị
        # tách nhầm thành hai lựa chọn.
        bad = [o for o in options if "," in o]
        if bad:
            raise ValueError(f"option của data validation không được chứa dấu phẩy: {bad}")
        dv = DataValidation(
            type="list", formula1='"' + ",".join(options) + '"', allow_blank=True
        )
    else:
        dv = DataValidation(
            type=kind,
            operator="between",
            formula1=str(rule.get("min", 0)),
            formula2=str(rule.get("max", 10**9)),
            allow_blank=True,
        )

    dv.error = "Giá trị không hợp lệ cho cột này."
    dv.errorTitle = "Sai định dạng"
    ws.add_data_validation(dv)
    dv.add(f"{letter}{first_row}:{letter}{last_row}")


def build_sheet(wb: Workbook, sheet: dict, first: bool) -> None:
    ws = wb.active if first else wb.create_sheet()
    ws.title = sheet["name"]

    columns = sheet["columns"]
    sample_rows = sheet["sampleRows"]
    blank_rows = sheet.get("blankRows", 20)

    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=col["header"])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = col.get("width", 16)

    ws.row_dimensions[1].height = 32

    first_data_row = 2
    last_data_row = first_data_row + len(sample_rows) + blank_rows - 1

    for offset in range(len(sample_rows) + blank_rows):
        row_num = first_data_row + offset
        values = sample_rows[offset] if offset < len(sample_rows) else {}

        for i, col in enumerate(columns, start=1):
            cell = ws.cell(row=row_num, column=i)
            cell.border = BORDER

            if col["type"] == "formula":
                # Dòng trống cũng cài sẵn công thức, để người dùng nhập tiếp
                # mà không phải tự kéo công thức xuống.
                cell.value = resolve_formula(col["formula"], columns, row_num)
                cell.fill = FORMULA_FILL
                fmt = col.get("format", "number")
            else:
                if col["key"] in values:
                    cell.value = values[col["key"]]
                fmt = col["type"]

            cell.number_format = NUMBER_FORMATS[fmt]

    for i, col in enumerate(columns, start=1):
        apply_validation(ws, col, get_column_letter(i), first_data_row, last_data_row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{last_data_row}"


def build_guide_sheet(wb: Workbook, spec: dict) -> None:
    """Sheet hướng dẫn kèm link về trang gốc.

    File .xlsx được chia sẻ lại qua Zalo/email tách rời khỏi website, nên nhúng
    link nguồn là cách giữ đường quay lại duy nhất.
    """
    ws = wb.create_sheet("Hướng dẫn")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100

    rows: list[tuple[str, Font | None]] = [(spec["h1"], TITLE_FONT), ("", None)]

    rows.append(("CÁCH DÙNG", Font(bold=True, size=12)))
    for i, step in enumerate(spec["howToUse"], start=1):
        rows.append((f"{i}. {step['step']}", Font(bold=True)))
        rows.append((f"   {step['detail']}", None))
    rows.append(("", None))

    rows.append(("CÔNG THỨC TRONG FILE", Font(bold=True, size=12)))
    for sheet in spec["sheets"]:
        for col in sheet["columns"]:
            if col.get("formula"):
                formula = resolve_formula(col["formula"], sheet["columns"], 2)
                rows.append((f"{col['header']}  —  {formula}", Font(bold=True)))
                rows.append((f"   {col['note']}", None))
    rows.append(("", None))

    url = f"{SITE_URL}/mau-excel/{spec['category']}/{spec['slug']}"
    rows.append(("Bản cập nhật mới nhất và hướng dẫn chi tiết:", Font(bold=True)))
    rows.append((url, Font(color="1D4ED8", underline="single")))

    for idx, (text, font) in enumerate(rows, start=1):
        cell = ws.cell(row=idx, column=2, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if font:
            cell.font = font

    ws.cell(row=len(rows), column=2).hyperlink = url


def build(spec_path: Path) -> Path:
    spec = json.loads(spec_path.read_text(encoding="utf-8"))

    wb = Workbook()
    for i, sheet in enumerate(spec["sheets"]):
        build_sheet(wb, sheet, first=(i == 0))
    build_guide_sheet(wb, spec)

    out_path = OUT_DIR / spec["category"] / f"{spec['slug']}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> int:
    specs = sorted(DATA_DIR.glob("*/*.json"))
    if not specs:
        print("Không tìm thấy spec nào trong data/templates/", file=sys.stderr)
        return 1

    for spec_path in specs:
        try:
            out = build(spec_path)
        except Exception as exc:  # noqa: BLE001
            print(f"✗ {spec_path.name}: {exc}", file=sys.stderr)
            return 1
        print(f"✓ {out.relative_to(ROOT)}  ({out.stat().st_size:,} bytes)")

    print(f"\nĐã sinh {len(specs)} file.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
